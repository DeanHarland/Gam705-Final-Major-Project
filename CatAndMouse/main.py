import pygame
import random
import math
import os
import sys
import wall
import character
import time
from openpyxl import workbook
from openpyxl import Workbook

import collections
import heapq

pygame.init()
# Open a window on screen
screenHeight = 900
screenWidth = 1200
screenWindow = pygame.display.set_mode([screenWidth, screenHeight])
screenWindow.fill((20, 90, 20))

# Title
pygame.display.set_caption("Cat And Mouse")

# Colour
black = (0,0,0)
dimGrey = (50,50,50)
dimGrey1 = (128,128,128)
greyFloor = (96,96,96)
brown = (102,34,0)
browner = (77, 34, 0)
brownerer = (51, 17, 0)
darkestBrown = (26, 9, 0)
darkGreen = (0,51,0)
blue = (51,153,255)
darkBlue = (51,153,200)
treeGreen = (0,153,0)
pink = (255,105,180)
red = (255,0,0)
# UI
font = pygame.font.Font('freesansbold.ttf', 26)
textX = 10
textY = 10
startTime = 0
# Timer
TIME = pygame.USEREVENT + 1
pygame.time.set_timer(TIME, 1000)
timer = 100

testClock = pygame.time.Clock()
dt = testClock.tick()
timeElaps = 0

# Player
playerX = screenWidth - 100
playerY = 100

# Spike ability
spikeImg = pygame.image.load("nail.png")
spikeX = 0
spikeY = playerY
spikeXSpeed = 0
spikeYSpeed = 0
spikeState = "ready"
spikePosSpeed = 5
spikeNegSpeed = -5

# Run away mousey
runAwayCharges = 1
canRunAway = True

# Game Over
gameOverFont = pygame.font.Font('freesansbold.ttf', 64)

# Data handling
wb = Workbook()
# Grab the active worksheet
ws = wb.active

ws['A1'] = "PlayerXPos"
ws['B1'] = "PlayerYPos"
ws['C1'] = "MouseXPos"
ws['D1'] = "MouseYPos"
ws['E1'] = "Time"
ws['F1'] = "SpikeActive"
ws['G1'] = "SpikeXPos"
ws['H1'] = "SpikeYPos"
ws['I1'] = "LeftPressed"
ws['J1'] = "RightPressed"
ws['K1'] = "UpPressed"
ws['L1'] = "DownPressed"

wb.save('CatData.csv')
# ws.append([playerX, playerY,mouseX,mouseY,time,spikeState,spikeX, spikeY, ])

def ShowGameOver(x,y):
    gameOverText = gameOverFont.render(("Game Over"),True, (0, 0, 0))
    screenWindow.blit(gameOverText,(x,y) )

class Player(pygame.sprite.Sprite):
    def __init__(self,x,y):

        super().__init__()

        self.playerXSpeed = 0
        self.playerYSpeed = 0
        self.walls = None

        self.image = pygame.image.load('Cat.png')
        self.rect = self.image.get_rect()
        self.rect.y = y
        self.rect.x = x

        self.isCollide = False

        #removable?
        self.canLeft = True
        self.canRight = True


    def drawPlayer(self,x,y):
        screenWindow.blit(self.image, (round(x), round(y)))


player = Player(30,30)
player.rect.x = playerX
player.rect.y = playerY
# Abilities
def shootSpike(x, y):
# Shoots a spike out infront of the players movement
    global spikeState
    spikeState = "fired"
    screenWindow.blit(spikeImg,(x,y))

# pounce

pounceLast = pygame.time.get_ticks()
pounceCooldown = 3000
pounceDistance = 100

def pounceAbility(x,y,direction):
    print("prepounce")
    leftPounce = x - pounceDistance
    rightPounce = x + pounceDistance
    upPounce = y - pounceDistance
    downPounce = y + pounceDistance
    global pounceLocationX
    pounceLocationX = 0
    global pounceLocationY
    pounceLocationY = 0
    if direction == "left":
        print("left start")
        pounceLocationX = x - pounceDistance
# Use current X + pounce distance to determine jump distance? but how to make it jump? speed up?
# Need sprint stamina for for loop?
# Which direction? how will it know?






class Mouse(pygame.sprite.Sprite):
    def __init__(self,x,y):

        super().__init__()

        self.mouseXSpeed = 0
        self.mouseYSpeed = 0
        self.walls = None

        self.image = pygame.image.load('Mouse.png')
        self.rect = self.image.get_rect()
        self.rect.y = y
        self.rect.x = x

        self.isCollide = False
    def drawMouse(self,x,y):
        screenWindow.blit(self.image, (round(x), round(y)))

mouseX = 50
mouseY = 850
mouse = Mouse(30,30)
mouse.rect.x = mouseX
mouse.rect.y = mouseY

class Generator(pygame.sprite.Sprite):
    def __init__(self,x,y):

        super().__init__()
        activated = False
        completed = False
        self.image = pygame.image.load('generator.png')
        self.rect = self.image.get_rect()
        self.rect.y = y
        self.rect.x = x


def drawGenerator(self,x, y):
    screenWindow.blit(self.image, (round(x), round(y)))

genX = 600
genY = 50
generator = Generator(30,30)
generator.rect.x = genX
generator.rect.y = genY
genActive = False
class Generator2(pygame.sprite.Sprite):
    def __init__(self,x,y):

        super().__init__()
        activated = False
        completed = False
        self.image = pygame.image.load('generator.png')
        self.rect = self.image.get_rect()
        self.rect.y = y
        self.rect.x = x


def drawGenerator2(self,x, y):
    screenWindow.blit(self.image, (round(x), round(y)))

gen2X = 1150
gen2Y = 775
generator2 = Generator2(30,30)
generator2.rect.x = gen2X
generator2.rect.y = gen2Y
gen2Active = False
# Game over/restart
gameOver = False

# Keyboard Variables
pressed_left = False
pressed_right = False
pressed_up = False
pressed_down = False
pressed_Q = False
pressed_W = False

clock = pygame.time.Clock()


DECWALL  = pygame.sprite.Group()
wall99 = wall.Wall((screenWidth/4-5),(screenHeight-415),30,210,black,screenWindow)
wall98 = wall.Wall((screenWidth/4-5),115,30,310,black,screenWindow)
wall97 = wall.Wall((screenWidth/4-5),20,30,50,black,screenWindow)

wall96 = wall.Wall(10,(screenHeight/4-5),105,30,black,screenWindow)
wall95 = wall.Wall((screenWidth/4-115),(screenHeight/4-5),110,30,black   ,screenWindow)

wall94 = wall.Wall((screenWidth/4*3-5),115,30,310,black,screenWindow)
wall93 = wall.Wall((screenWidth/4*3-5),20,30,55,black,screenWindow)
# right side boot
wall92 = wall.Wall((screenWidth/4*3-5),495,30,310,black,screenWindow)

wall91 = wall.Wall((screenWidth/4*3+20),(screenHeight/3-5),100,30,black,screenWindow)
wall90 = wall.Wall((screenWidth-110),(screenHeight/3-5),100,30,black,screenWindow)
wall89 = wall.Wall((screenWidth/4*3+20),(screenHeight/4*3-5),100,30,black,screenWindow)
wall88 = wall.Wall((screenWidth-110),(screenHeight/4*3-5),100,30,black,screenWindow)

# MID SIDE?
wall87 = wall.Wall((screenWidth/4+55),screenHeight/4-5,(screenWidth/2-40),30,black,screenWindow)  # top mid
wall86 = wall.Wall((screenWidth/4 ),screenHeight/4*3-5,(screenWidth/2),30,black,screenWindow)  # bottom mid

# Fountain
wall85 = wall.Wall(screenWidth/2-80 ,screenHeight/2+45,160,30,black,screenWindow)
wall84 = wall.Wall(screenWidth/2-80 ,screenHeight/2-55,160,30,black,screenWindow)

wall83 = wall.Wall(screenWidth/2+40 ,screenHeight/2-50,30,120,black,screenWindow)
wall82 = wall.Wall(screenWidth/2-70,screenHeight/2-50,30,120,black,screenWindow)

wall81 = wall.Wall(screenWidth/2-50,screenHeight/2-50,100,100,blue,screenWindow)
# Bench
wall80 = wall.Wall(screenWidth/2+65 ,screenHeight/2-25,20,70,brown,screenWindow)
wall79 = wall.Wall(screenWidth/2-85 ,screenHeight/2-25,20,70,brown,screenWindow)

# Tree
wall78 = wall.Wall(1100 ,200,70,70,black,screenWindow)
wall77 = wall.Wall(1000 ,50,70,70,black,screenWindow)
wall76 = wall.Wall(950 ,190,50,50,black,screenWindow)

# Upper Mid
wall75 = wall.Wall((screenWidth/4+195),20,30,75,black,screenWindow) # topleft
wall74 = wall.Wall((screenWidth/4+395),20,30,75,black,screenWindow) # topright

wall73 = wall.Wall((screenWidth/4+195),145,30,75,black,screenWindow) # BL
wall72 = wall.Wall((screenWidth/4+395),145,30,75,black,screenWindow) #BR


# bottom fences
wall71 = wall.Wall(395,745,25,110,black,screenWindow) # L1
wall70 = wall.Wall(495,745,25,110,black,screenWindow)
wall69 = wall.Wall(695,745,25,110,black,screenWindow)
wall68 = wall.Wall(795,745,25,110,black,screenWindow)

# Middle left shop
wall67 = wall.Wall(5,345,160,210,black,screenWindow)
wall66 = wall.Wall(225,345,20,20,black,screenWindow)
wall65 = wall.Wall(225,535,20,20,black,screenWindow)

#house
wall64 = wall.Wall((screenWidth/4-5),screenHeight/4*3-5,35,110,black,screenWindow)
wall63 = wall.Wall((screenWidth/4-5),screenHeight-75,35,70,black,screenWindow)
wall62 = wall.Wall(5,(screenHeight/4*3)-5,110,35,black,screenWindow)
wall61 = wall.Wall((screenWidth/4-110)-5,(screenHeight/4*3)-5,120,35,black,screenWindow)

wall60 = wall.Wall((screenWidth/4-100)-5,(screenHeight/4*3+90)-5,35,90,black,screenWindow)

DECWALL.add(wall99,wall80,wall79,wall81,wall98,wall97, wall96,wall95,wall94,wall93,wall92,wall91,wall90,wall89,wall88,wall87,wall86,wall85,wall84,wall83,wall82, wall78,wall77, wall76,wall75,wall74,wall73,wall72,wall71,wall70,
            wall69,wall68, wall67,wall66, wall65,wall64,wall63,wall62,wall61,wall60)



# Walls - created a group and add each instance of wall to it.
WALLS = pygame.sprite.Group()

# X,Y,Height,Width,Height,Colour,Screen.
wall1 = wall.Wall(0,0,screenWidth,20,black,screenWindow)                    # top wall
wall2 = wall.Wall(0,(screenHeight-10),screenWidth,20,black,screenWindow)        # bot wall
wall3 = wall.Wall(0,0,10,screenHeight,black,screenWindow)                       # Left Wall
wall4 = wall.Wall((screenWidth-10),0,20,screenHeight,black,screenWindow)        # right wall
# LEFT SIDE
wall5 = wall.Wall((screenWidth/4),(screenHeight-410),20,200,dimGrey,screenWindow)
wall16 = wall.Wall((screenWidth/4),120,20,300,dimGrey,screenWindow)
wall15 = wall.Wall((screenWidth/4),20,20,50,dimGrey,screenWindow)
# house home
wall25 = wall.Wall((screenWidth/4),screenHeight/4*3,25,100,brown,screenWindow)
wall26 = wall.Wall((screenWidth/4),screenHeight-70,25,60,brown,screenWindow)
wall22 = wall.Wall(10,(screenHeight/4*3),100,25,brown,screenWindow)
wall24 = wall.Wall((screenWidth/4-110),(screenHeight/4*3),110,25,brown,screenWindow)
wall47 = wall.Wall((screenWidth/4-100),(screenHeight/4*3+90),25,80,brown,screenWindow)
# left left side
wall21 = wall.Wall(10,(screenHeight/4),100,20,dimGrey,screenWindow)
wall23 = wall.Wall((screenWidth/4-110),(screenHeight/4),110,20,dimGrey,screenWindow)
# RIGHT SIDE TOP
wall10 = wall.Wall((screenWidth/4*3),120,20,300,dimGrey,screenWindow)
wall11 = wall.Wall((screenWidth/4*3),20,20,50,dimGrey,screenWindow)
# right side boot
wall12 = wall.Wall((screenWidth/4*3),500,20,300,dimGrey,screenWindow)
# right right side
wall17 = wall.Wall((screenWidth/4*3+20),(screenHeight/3),100,20,dimGrey,screenWindow)
wall18 = wall.Wall((screenWidth-110),(screenHeight/3),100,20,dimGrey,screenWindow)
wall19 = wall.Wall((screenWidth/4*3+20),(screenHeight/4*3),100,20,dimGrey,screenWindow)
wall20 = wall.Wall((screenWidth-110),(screenHeight/4*3),100,20,dimGrey,screenWindow)
# MID SIDE?
wall13 = wall.Wall((screenWidth/4+60),screenHeight/4,(screenWidth/2-40),20,dimGrey,screenWindow)  # top mid
wall14 = wall.Wall((screenWidth/4 ),screenHeight/4*3,270,20,dimGrey,screenWindow)  # bottom mid
wall46 = wall.Wall((screenWidth/4+330 ),screenHeight/4*3,270,20,dimGrey,screenWindow)

# Fountain
wall27 = wall.Wall(screenWidth/2-75 ,screenHeight/2+50,150,20,dimGrey1,screenWindow)
wall28 = wall.Wall(screenWidth/2-75 ,screenHeight/2-50,150,20,dimGrey1,screenWindow)

wall29 = wall.Wall(screenWidth/2+45 ,screenHeight/2-50,20,100,dimGrey1,screenWindow)
wall30 = wall.Wall(screenWidth/2-65,screenHeight/2-50,20,100,dimGrey1,screenWindow)

#Tree
wall31 = wall.Wall(1105 ,205,60,60,treeGreen,screenWindow)
wall32 = wall.Wall(1005 ,55,60,60,treeGreen,screenWindow)

wall33 = wall.Wall(955 ,195,40,40,treeGreen,screenWindow)

# Upper Mid
wall34 = wall.Wall((screenWidth/4+200),20,20,70,dimGrey,screenWindow) # topleft
wall35 = wall.Wall((screenWidth/4+400),20,20,70,dimGrey,screenWindow) # topright

wall36 = wall.Wall((screenWidth/4+200),150,20,70,dimGrey,screenWindow) # BL
wall37 = wall.Wall((screenWidth/4+400),150,20,70,dimGrey,screenWindow) #BR

# Fences
wall38 = wall.Wall(400,750,15,100,darkestBrown,screenWindow)
wall39 = wall.Wall(500,750,15,100,darkestBrown,screenWindow)
wall40 = wall.Wall(700,750,15,100,darkestBrown,screenWindow)
wall41 = wall.Wall(800,750,15,100,darkestBrown,screenWindow)


# shop
wall42 = wall.Wall(10,350,150,200,dimGrey1,screenWindow)
wall43 = wall.Wall(230,350,10,10,dimGrey1,screenWindow)
wall44 = wall.Wall(230,540,10,10,dimGrey1,screenWindow)
wall45 = wall.Wall(150,450,20,70,black,screenWindow)

# 46 used
# 47used

WALLS.add(wall1,wall2,wall3,wall4,wall5,wall14,wall15,wall25,wall26,wall22,wall24,wall28,wall21,wall23,wall10,wall11, wall12, wall17,wall18,wall19,wall20,wall13,wall16, wall27,wall28,wall29,wall30,wall31,wall32, wall33,wall34
          ,wall35,wall36,wall37,wall38,wall39,wall40,wall41, wall42,wall43,wall44,wall45,wall46,wall47)
player.walls = WALLS

TESTWALLS = []
# Width ones
#'''
def makeWallNodes(x,y,width,height):
    startX = round(x/30)
    startY = round(y/30)
    loopCountW = round(width/30)
    i = 0
    while i <= loopCountW:

        #startX += 1
        nodeXY = (startX, startY)
        TESTWALLS.append(nodeXY)
        #print(nodeXY)
        startX += 1
        i+= 1

#'''
def makeHeightWallNodes(x,y,width,height):
    startX = round(x/30)
    startY = round(y/30)
    loopCountH = round(height/30)
    i = 0
    while i <= loopCountH:

        #startY += 1
        nodeXY = (startX, startY)
        TESTWALLS.append(nodeXY)
        #print(nodeXY)
        startY += 1
        i+= 1
for wall in WALLS:
    makeWallNodes(wall.x,wall.y,wall.width,wall.height)

# '''
for wall in WALLS:
    #makeWallNodes(wall.x,wall.y,wall.width,wall.height)
    makeHeightWallNodes(wall.x, wall.y, wall.width, wall.height)
# '''



def drawFloors():
    # pygame.draw.rect(screenWindow,dimGrey,[10,680,300,420])
    pygame.draw.rect(screenWindow, browner, [10, 680, 30, 300])
    pygame.draw.rect(screenWindow, brownerer, [40, 680, 30, 300])
    pygame.draw.rect(screenWindow, browner, [70, 680, 30, 300])
    pygame.draw.rect(screenWindow, brownerer, [100, 680, 30, 300])
    pygame.draw.rect(screenWindow, browner, [130, 680, 30, 300])
    pygame.draw.rect(screenWindow, brownerer, [160, 680, 30, 300])
    pygame.draw.rect(screenWindow, browner, [190, 680, 30, 300])
    pygame.draw.rect(screenWindow, brownerer, [220, 680, 30, 300])
    pygame.draw.rect(screenWindow, browner, [250, 680, 30, 300])
    pygame.draw.rect(screenWindow, brownerer, [280, 680, 30, 300])
    # Monster Spawn Floor
    pygame.draw.rect(screenWindow,darkGreen,[900,0,300,300])
    pygame.draw.rect(screenWindow, darkGreen, [1000, 200, 100, 280])

    # Middle Area
    pygame.draw.rect(screenWindow, greyFloor, [300, 580, 600, 100]) # B
    pygame.draw.rect(screenWindow, greyFloor, [300, 240, 600, 100]) # T
    pygame.draw.rect(screenWindow, greyFloor, [300, 240, 100, 400]) # L
    pygame.draw.rect(screenWindow, greyFloor, [800, 240, 100, 400]) # R
    pygame.draw.rect(screenWindow, greyFloor, [500, 0, 220, 220])

    pygame.draw.rect(screenWindow, greyFloor, [1100, 725, 200, 150])
    pygame.draw.rect(screenWindow, dimGrey, [1140, 765, 50, 50])
    # flowers?

    #generator lights
    pygame.draw.rect(screenWindow, red, [genX, genY - 5, 25, 10])
    if genActive == True:
        pygame.draw.rect(screenWindow, treeGreen, [genX, genY - 5, 25, 10])

    pygame.draw.rect(screenWindow, red, [gen2X, gen2Y - 5, 25, 10])
    if gen2Active == True:
        pygame.draw.rect(screenWindow, treeGreen, [gen2X, gen2Y - 5, 25, 10])

    # Shop
    pygame.draw.rect(screenWindow, (20, 69, 20), [150, 350, 160, 200])
    #20, 90, 20

def checkForWalls():

    for wall in WALLS:

        if player.rect.colliderect(wall):
            player.isCollide = True
            break
        if not player.rect.colliderect(wall):
            player.isCollide = False


def showTime(x,y):
    timeText = font.render("Time: " + str(timer),True,(0,0,0))
    screenWindow.blit(timeText,(x,y))

# A Star

class PriorityQueue:
    def __init__(self):
        self.elements = []

    def empty(self):
        return len(self.elements) == 0

    def put(self, item, priority):
        heapq.heappush(self.elements, (priority, item))

    def get(self):
        return heapq.heappop(self.elements)[1]

class SquareGrid:
    def __init__(self, width, height):
        self.width = width
        self.height = height
        self.walls = []

    def in_bounds(self, id):
        (x, y) = id
        return 0 <= x < self.width and 0 <= y < self.height

    def passable(self, id):
        return id not in self.walls

    def neighbors(self, id):
        (x, y) = id
        results = [(x + 1, y), (x, y - 1), (x - 1, y), (x, y + 1),(x + 1, y + 1), (x+1,y-1), (x-1,y-1),(x-1, y+1) ]
        if (x + y) % 2 == 0: results.reverse()  # aesthetics
        results = filter(self.in_bounds, results)
        results = filter(self.passable, results)
        return results

class GridWithWeights(SquareGrid):
    def __init__(self, width, height):
        super().__init__(width, height)
        self.weights = {}

    def cost(self, from_node, to_node):
        return self.weights.get(to_node, 1)

#NODEWALLS = [(8,82)]
################################################

grid = GridWithWeights(40, 30 )

###############################################
grid.walls = TESTWALLS #NODEWALLS

def heuristic(a, b):
    (x1, y1) = a
    (x2, y2) = b
    return abs(x1 - x2) + abs(y1 - y2)

# test
def a_star_search(graph, start, goal):
    frontier = PriorityQueue()
    frontier.put(start, 0)
    came_from = {}
    cost_so_far = {}
    came_from[start] = None
    cost_so_far[start] = 0

    while not frontier.empty():
        current = frontier.get()

        if current == goal:

            break

        for next in graph.neighbors(current):
            new_cost = cost_so_far[current] + graph.cost(current, next)
            if next not in cost_so_far or new_cost < cost_so_far[next]:
                cost_so_far[next] = new_cost
                priority = new_cost + heuristic(goal, next)
                frontier.put(next, priority)
                came_from[next] = current


    return came_from, cost_so_far


#start = ((round(playerX/10)), round(playerY/10))

#goal = ((round(playerX / 10)), (round(playerY / 10)))
start = ((round(mouseX / 30)), round(mouseY / 30))
goal = ((round(genX / 30)), (round(genY / 30)))
came_from, cost_so_far = a_star_search(grid, start,goal)

def reconstruct_path(came_from, start, goal):
    current = goal
    path = []
    while current != start:
        path.append(current)
        current = came_from[current]
    path.append(start) # optional
    path.reverse() # optional
    return path

path =  reconstruct_path(came_from,start,goal)

firstNode = path[1]

def checkDistance(playerX, playerY, mouseX, mouseY):
    distance = math.sqrt((math.pow(playerX - mouseX, 2)) + (math.pow(playerY- mouseY, 2)))
    if distance < 200:
        return True
    else:
        return False


# # # Main game loop # # #
running = True
while running:


    screenWindow.fill((30, 90, 30))
    drawFloors()

    start = ((round(mouseX / 30)), round(mouseY / 30))
    mouseXNode = round(mouseX / 30)
    mouseYNode = round(mouseY / 30)
    start = ((round(mouseX / 30)), round(mouseY / 30))
    came_from, cost_so_far = a_star_search(grid, start, goal)
    a_star_search(grid,start,goal)
    path = reconstruct_path(came_from, start, goal)

    firstNode = path[1]
    firstNodeX, firstNodeY = firstNode

    if mouseXNode < firstNodeX:
        mouseX += 2
    elif mouseXNode > firstNodeX:
        mouseX -= 2

    if mouseYNode < firstNodeY:
        mouseY += 2
    elif mouseYNode > firstNodeY:
        mouseY -= 2

    if mouseXNode == firstNodeX:
        #print("X hit")
        pass
    if mouseYNode == firstNodeY:
        #print("Y hit")
        pass

    if mouse.rect.colliderect(generator):
        print("mouse hit gen")
        goal = ((round(gen2X / 30)), (round(gen2Y / 30)))
        start = ((round(mouseX / 30)), round(mouseY / 30))
        came_from, cost_so_far = a_star_search(grid, start, goal)
        genActive = True


    if mouse.rect.colliderect(generator2):
        print("mouse hit gen2")
        goal = ((round(genX / 30)), (round(genY / 30)))
        start = ((round(mouseX / 30)), round(mouseY / 30))
        came_from, cost_so_far = a_star_search(grid, start, goal)
        gen2Active = True

    # Move then remove from array? or just redo a star all the time so its always moving to first node

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

        elif event.type == pygame.KEYDOWN:  # check for key presses
            if event.key == pygame.K_LEFT:
                pressed_left = True
                player.playerXSpeed = 3
            elif event.key == pygame.K_RIGHT:
                pressed_right = True
                player.playerXSpeed = 3
            elif event.key == pygame.K_UP:
                pressed_up = True
                player.playerYSpeed = 3
            elif event.key == pygame.K_DOWN:
                pressed_down = True
                player.playerYSpeed = 3
            elif event.key == pygame.K_q:
                pressed_Q = True
                if spikeState == "ready":
                    spikeX = playerX
                    spikeY = playerY
                    if pressed_left == True:
                        if pressed_up == True:
                            spikeYSpeed = (spikeNegSpeed+2)
                            spikeY -= 15
                            spikeXSpeed = (spikeNegSpeed+2)
                            spikeX -= 15
                        if pressed_down == True:
                            spikeXSpeed = (spikeNegSpeed+1)
                            spikeX -= 15
                            spikeYSpeed = (spikePosSpeed-1)
                            spikeY += 15
                        else:
                            spikeXSpeed = spikeNegSpeed
                            spikeX -= 15
                    elif pressed_right == True:
                        if pressed_up == True:
                            spikeYSpeed = (spikeNegSpeed+2)
                            spikeY -= 15
                            spikeXSpeed = (spikePosSpeed-2)
                            spikeX += 15
                        if pressed_down == True:
                            spikeXSpeed = (spikePosSpeed-1)
                            spikeX += 15
                            spikeYSpeed = (spikePosSpeed-1)
                            spikeY += 15
                        else:
                            spikeXSpeed = spikePosSpeed
                            spikeX += 15
                    elif pressed_up == True:
                        spikeYSpeed = spikeNegSpeed
                        spikeY -= 15
                    elif pressed_down == True:
                        spikeYSpeed = spikePosSpeed
                        spikeY += 15
                    else:
                        spikeYSpeed = 4
                    shootSpike(spikeX, spikeY)
                    # Pounce
            elif event.key == pygame.K_w:
                pressed_W = True







                now = pygame.time.get_ticks()
                if now - pounceLast >= pounceCooldown:
                    pounceLast = now


                    #player.playerXSpeed = 5
                            #pounceAbility(playerX, playerY, "left")
                            #playerX = pounceLocationX


        elif event.type == TIME:
            timer -= 1
            ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])


        elif event.type == pygame.KEYUP:  # check for key releases
            if event.key == pygame.K_LEFT:
                pressed_left = False
                player.playerXSpeed = 3
            elif event.key == pygame.K_RIGHT:
                pressed_right = False
                player.playerXSpeed = 3
            elif event.key == pygame.K_UP:
                pressed_up = False
                player.playerYSpeed = 3
            elif event.key == pygame.K_DOWN:
                pressed_down = False
                player.playerYSpeed = 3
            elif event.key == pygame.K_q:
                pressed_Q = False
            elif event.key == pygame.K_w:
                pressed_W = False
    '''
    # maybe get direction by using velocity if last X > new X, -X
    if time >= 0:

        if mouseX < gen2X:
            mouseX += 2


        if playerX <= (mouseX+60):
            mouseX -= 5
        if  playerX <= (mouseX-60):
            mouseX += 2
    '''


    if pressed_left:
        playerX -= player.playerXSpeed
        player.rect.x = playerX
        player.rect.y = playerY
        checkForWalls()
        ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])
        if player.isCollide == True:
            playerX += player.playerXSpeed

    if pressed_right:
        playerX += player.playerXSpeed
        player.rect.x = playerX
        player.rect.y = playerY
        checkForWalls()
        ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])
        if player.isCollide == True:
            playerX -= player.playerXSpeed

    if pressed_up:
        playerY -= player.playerYSpeed
        player.rect.x = playerX
        player.rect.y = playerY
        checkForWalls()
        ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])
        if player.isCollide == True:
            playerY += player.playerYSpeed

    if pressed_down:
        playerY += player.playerYSpeed
        player.rect.x = playerX
        player.rect.y = playerY
        checkForWalls()
        ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])
        if player.isCollide == True:
            playerY -= player.playerYSpeed

    if pressed_W:
        ws.append([playerX, playerY, mouseX, mouseY, timer, spikeState, spikeX, spikeY,pressed_left,pressed_right,pressed_up,pressed_down])
        pass

    spikeCollide = False
    spikeRect = pygame.Rect(spikeX, spikeY, 30, 30)
    if spikeState == "fired":
        shootSpike(spikeX, spikeY)
        spikeX += spikeXSpeed
        spikeY += spikeYSpeed
        spikeRect = pygame.Rect(spikeX, spikeY,30,30)
        for wall in WALLS:
            if spikeRect.colliderect(wall):
                spikeCollide = True
                spikeXSpeed = 0
                spikeYSpeed = 0
                spikeState = "ready"
                break
            if not spikeRect.colliderect(wall):
                spikeCollide = False

    if player.rect.colliderect(generator):
        genActive = False
    if player.rect.colliderect(generator2):
        gen2Active = False

    if player.rect.colliderect(mouse):
        gameOver = True
    if mouse.rect.colliderect(spikeRect):
        gameOver = True

    #checkDistance(playerX,playerY,mouseX,mouseY)
    nearEachOther = checkDistance(playerX,playerY,mouseX,mouseY)
    if nearEachOther:
        if runAwayCharges == 1:
            if playerX > mouseX:
                goal = ((round(genX / 30)), (round(genY / 30)))
                start = ((round(mouseX / 30)), round(mouseY / 30))
                came_from, cost_so_far = a_star_search(grid, start, goal)
                runAwayCharges =- 1
            elif playerX < mouseX:
                goal = ((round(gen2X / 30)), (round(gen2Y / 30)))
                start = ((round(mouseX / 30)), round(mouseY / 30))
                came_from, cost_so_far = a_star_search(grid, start, goal)
                runAwayCharges = - 1





    if timer == 0:
        gameOver = True
    if timer == 90:
        runAwayCharges = 1
    if timer == 80:
        runAwayCharges = 1
    if timer == 70:
        runAwayCharges = 1
    if timer == 60:
        runAwayCharges = 1

    if gameOver == True:
       # time = 0
        ShowGameOver(450,300)
        wb.save('CatData.csv')
        mouseX = 0
        mouseY = 0
        time.sleep(1)
        exit()
        pass


    drawGenerator(generator, genX, genY)
    drawGenerator(generator, gen2X,gen2Y)
    player.drawPlayer(playerX,playerY)
    Mouse.drawMouse(mouse, mouseX, mouseY)
    player.rect.x = playerX
    player.rect.y = playerY
    mouse.rect.x = mouseX
    mouse.rect.y = mouseY
    for wall in DECWALL:
        wall.draw()
    for wall in WALLS:
        wall.draw()
    showTime(30,30)

    pygame.display.update()
    clock.tick(60)